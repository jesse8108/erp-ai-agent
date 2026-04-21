"""
根据用户人工标注过的 duplicates.xlsx 合并重复公司 + 清理垃圾数据 + 转嫁简称。

输入:
  - /home/claude/work/companies.xlsx      (1712 家原始清洗数据)
  - /home/claude/work/aliases.xlsx        (3470 条自动生成简称)
  - /mnt/user-data/uploads/duplicates.xlsx (用户标注过的版本)

规则:
  - 每组保留用户在 duplicates.xlsx 里留下的那一行
  - 被删公司的简称转嫁给保留的, 去重
  - 组 11 (珠海瑞祥丰) 合并为 1 条, 标注已倒闭
  - 组 19 (国贸启润) 不合并, 两条都保留
  - 额外清理 11 条个人名占位数据

输出:
  - companies_final.xlsx
  - aliases_final.xlsx
  - cleanup_report.md
"""
from openpyxl import load_workbook, Workbook
from collections import defaultdict

INPUT_COMPANIES = "/home/claude/work/companies.xlsx"
INPUT_ALIASES = "/home/claude/work/aliases.xlsx"
INPUT_DUPLICATES = "/mnt/user-data/uploads/duplicates.xlsx"

# 第一类垃圾数据: 个人名占位
GARBAGE_NAMES = {
    "董总", "损毁单位",
    "徐正", "徐惠忠", "孙其军", "周涛", "济宁刘",
    "闫晓景", "刘文超", "尹东", "吴泳谊",
}

# 组 11 (珠海瑞祥丰) 特殊处理
GROUP_11_KEEP = "珠海市瑞祥丰化工新材料有限公司"

# 原 22 组每组涉及的所有公司 (从 detect_duplicates.py 的输出)
ORIGINAL_GROUPS = {
    1: ["山东九塑科技发展有限公司--不用", "（弃用）山东九塑科技发展有限公司", "山东九塑新材料科技股份有限公司"],
    2: ["中科佳一（嘉兴）新材料科技有限公司", "中科佳一(嘉兴)新材料科技有限公司"],
    3: ["山东悦齐新材料科技有限公司", "淄博悦齐新材料科技有限公司"],
    4: ["上海赫诚运材料科技有限公司", "上海赫诚运能源有限公司"],
    5: ["沧州德福塑料包装有限公司", "沧州企行塑料包装有限公司"],
    6: ["班森实业（深圳）有限公司", "班森实业(深圳)有限公司"],
    7: ["江西乾财商贸有限公司", "江西乾财工贸有限公司"],
    8: ["滨州三星包装科技有限公司", "滨州三星工贸有限公司"],
    9: ["河南晟翔浩峻贸易有限公司", "河南晨翔浩峻贸易有限公司"],
    10: ["嘉创新型材料（廊坊）有限公司", "嘉创新型材料(廊坊)有限公司"],
    11: ["珠海市瑞祥丰化工新材料有限公司", "珠海瑞祥丰化工新材料有限公司"],
    12: ["致远顺材料科技(苏州)有限公司", "致远顺材料科技（苏州）有限公司"],
    13: ["苏州日益鑫科技贸易有限公司", "州日益鑫科技贸易有限公司"],
    14: ["浙江景诚实业有限公 司", "浙江景诚实业有限公司"],
    15: ["东莞市塑之源塑胶原料有限公司", "东莞塑之源塑胶原料有限公司"],
    16: ["东莞市辰玉新材料科技有限公司", "东莞辰玉新材料科技有限公司"],
    17: ["福建华航发展有限公司", "福建华航有限公司"],
    18: ["石图里包装材料（扬州）有限公司", "石图里包装材料(扬州)有限公司"],
    19: ["国贸启润(上海)有限公司", "国贸启润（杭州）有限公司"],
    20: ["万聚荣华进出口贸易（广州）有限公司", "万聚荣华进出口贸易(广州)有限公司"],
    21: ["厦门国贸集团股份有限公司", "厦门国贸化工有限公司"],
    22: ["苏州金誉致化工新材料科技有限公司", "苏州金誉致化工新材料有限公司"],
}


def load_user_decisions():
    """读用户编辑过的 duplicates.xlsx, 得到每组保留哪些公司。"""
    wb = load_workbook(INPUT_DUPLICATES, read_only=True)
    ws = wb.active
    keep_by_group = defaultdict(list)
    for row in ws.iter_rows(values_only=True):
        if row[0] is None or row[0] == "组号":
            continue
        group = row[0]
        name = (row[4] or "").strip() if row[4] else ""
        if name:
            keep_by_group[group].append(name)
    return dict(keep_by_group)


def compute_merges(keep_by_group):
    """推导合并映射。返回 (merge_map, mark_inactive)."""
    merge_map = {}        # 被删 -> 保留
    mark_inactive = set() # 要标注 is_active=false 的

    for group, originals in ORIGINAL_GROUPS.items():
        kept = keep_by_group.get(group, [])
        if group == 19:
            continue  # 不合并
        if group == 11:
            keeper = GROUP_11_KEEP
            mark_inactive.add(keeper)
            for orig in originals:
                if orig != keeper:
                    merge_map[orig] = keeper
            continue
        if not kept:
            continue
        keeper = kept[0]
        for orig in originals:
            if orig != keeper:
                merge_map[orig] = keeper
    return merge_map, mark_inactive


def process_companies(merge_map, mark_inactive):
    """处理 companies.xlsx: 删重复 + 删垃圾 + 标注注销."""
    wb = load_workbook(INPUT_COMPANIES, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0]) + ["is_active", "business_status_override"]

    kept = []
    deleted = {}  # name -> reason

    for row in rows[1:]:
        name = (row[0] or "").strip() if row[0] else ""
        if not name:
            continue
        if name in GARBAGE_NAMES:
            deleted[name] = "垃圾数据"
            continue
        if name in merge_map:
            deleted[name] = f"合并到 {merge_map[name]}"
            continue
        is_active = name not in mark_inactive
        status_override = "已倒闭" if name in mark_inactive else ""
        kept.append(list(row) + [is_active, status_override])

    return header, kept, deleted


def process_aliases(merge_map, deleted_company_names):
    """转嫁简称 + 删垃圾简称 + 去重 + 过滤噪音."""
    wb = load_workbook(INPUT_ALIASES, read_only=True)
    ws = wb.active
    pairs = list(ws.iter_rows(values_only=True))[1:]

    # 噪音简称的特征: 包含这些子串的简称都过滤掉
    NOISE_MARKERS = ["--不用", "弃用", "(弃用)", "（弃用）"]

    def is_noise(alias: str) -> bool:
        """判断一个简称是否是从脏数据转嫁来的噪音。"""
        if not alias:
            return True
        # 包含噪音标记
        for marker in NOISE_MARKERS:
            if marker in alias:
                return True
        # 包含空格 (如 '景诚实业有限公 司')
        if " " in alias or "\u3000" in alias:  # 半角 + 全角空格
            return True
        # 包含 '有限公司' 全称 (不是真正的简称)
        if "有限公司" in alias or "有限责任公司" in alias:
            return True
        return False

    by_company = defaultdict(set)
    for formal, alias in pairs:
        if formal and alias:
            by_company[formal].add(alias)

    transferred = 0
    # 转嫁
    for src, tgt in merge_map.items():
        if src in by_company:
            src_aliases = by_company[src]
            tgt_aliases = by_company[tgt]
            # 转嫁时过滤噪音
            clean_new = {a for a in src_aliases if not is_noise(a)}
            new_ones = clean_new - tgt_aliases
            tgt_aliases.update(clean_new)
            transferred += len(new_ones)
            del by_company[src]

    # 删垃圾
    for g in GARBAGE_NAMES:
        if g in by_company:
            del by_company[g]

    # 防御性: 删其他被删公司名下残留
    for n in list(by_company.keys()):
        if n in deleted_company_names and n not in merge_map:
            del by_company[n]

    # 最后对所有保留公司的简称再过一遍噪音过滤
    # (防止原本就在主表里的公司简称里有噪音)
    noise_removed = 0
    for formal, aliases in by_company.items():
        clean = {a for a in aliases if not is_noise(a)}
        removed = len(aliases) - len(clean)
        if removed > 0:
            noise_removed += removed
            by_company[formal] = clean

    result = []
    for formal in sorted(by_company.keys()):
        for alias in sorted(by_company[formal], key=len):
            result.append((formal, alias))

    print(f"    噪音简称过滤掉 {noise_removed} 条")
    return result, transferred


def main():
    print("读取用户编辑过的 duplicates.xlsx ...")
    keep_by_group = load_user_decisions()
    print(f"  用户决定: {len(keep_by_group)} 组")

    print("\n推导合并映射 ...")
    merge_map, mark_inactive = compute_merges(keep_by_group)
    print(f"  待合并: {len(merge_map)} 家")
    print(f"  注销标记: {len(mark_inactive)} 家")

    print("\n处理 companies.xlsx ...")
    header, kept_companies, deleted_companies = process_companies(merge_map, mark_inactive)
    merge_count = sum(1 for r in deleted_companies.values() if "合并到" in r)
    garbage_count = len(deleted_companies) - merge_count
    print(f"  保留: {len(kept_companies)} 家")
    print(f"  删除: {len(deleted_companies)} 家 (合并 {merge_count} + 垃圾 {garbage_count})")

    print("\n处理 aliases.xlsx ...")
    final_aliases, transferred = process_aliases(merge_map, set(deleted_companies.keys()))
    print(f"  最终简称数: {len(final_aliases)}")
    print(f"  转嫁的简称: {transferred} 条")

    # ============================================================
    # 写文件
    # ============================================================
    wb_c = Workbook()
    ws_c = wb_c.active
    ws_c.title = "companies"
    ws_c.append(header)
    for row in kept_companies:
        ws_c.append(row)
    wb_c.save("/home/claude/work/companies_final.xlsx")

    wb_a = Workbook()
    ws_a = wb_a.active
    ws_a.title = "aliases"
    ws_a.append(["正式名", "简称"])
    for formal, alias in final_aliases:
        ws_a.append([formal, alias])
    wb_a.save("/home/claude/work/aliases_final.xlsx")

    with open("/home/claude/work/cleanup_report.md", "w", encoding="utf-8") as f:
        f.write("# 数据清理报告\n\n")
        f.write("## 摘要\n\n")
        f.write(f"- 原始公司数: 1712\n")
        f.write(f"- 最终公司数: {len(kept_companies)}\n")
        f.write(f"- 删除公司数: {len(deleted_companies)}\n")
        f.write(f"  - 合并重复: {merge_count} 条\n")
        f.write(f"  - 垃圾数据: {garbage_count} 条\n")
        f.write(f"- 标注注销: {len(mark_inactive)} 条\n")
        f.write(f"- 原始简称数: 3470\n")
        f.write(f"- 最终简称数: {len(final_aliases)}\n")
        f.write(f"- 转嫁的简称: {transferred} 条\n\n")

        f.write("## 合并明细\n\n")
        f.write("| 组号 | 保留 | 合并/删除 | 备注 |\n")
        f.write("|------|------|----------|------|\n")
        for group in sorted(ORIGINAL_GROUPS.keys()):
            originals = ORIGINAL_GROUPS[group]
            kept = keep_by_group.get(group, [])
            if group == 19:
                f.write(f"| 19 | {' / '.join(kept)} | — | **不合并**, 用户判断为不同公司 |\n")
            elif group == 11:
                merged = [o for o in originals if o != GROUP_11_KEEP]
                f.write(f"| 11 | {GROUP_11_KEEP} | {', '.join(merged)} | **已倒闭**, is_active=false |\n")
            else:
                keeper = kept[0] if kept else "(无)"
                merged = [o for o in originals if o != keeper]
                f.write(f"| {group} | {keeper} | {', '.join(merged) if merged else '—'} | |\n")

        f.write("\n## 清理的垃圾数据 (11 条)\n\n")
        for name in sorted(GARBAGE_NAMES):
            f.write(f"- {name}\n")

        f.write("\n## 关于 is_active / business_status_override\n\n")
        f.write("- 这两列是清理脚本新加的\n")
        f.write("- 导入数据库时填入 companies 表的 `is_active` 和 `business_status` 字段\n")
        f.write("- 绝大多数公司 `is_active=true`, `business_status_override=''`\n")
        f.write(f"- 只有组 11 (珠海瑞祥丰) 标为 `is_active=false`, `business_status_override='已倒闭'`\n\n")
        f.write("## 下一步\n\n")
        f.write("- 把 companies_final.xlsx 和 aliases_final.xlsx 放到 data-import/ 目录\n")
        f.write("- W1 建库完成后用导入脚本一次性入库\n")
        f.write("- companies 表的 `is_active=false` 记录会保留但不出现在业务员的可选列表里\n")

    print("\n✅ 完成")
    print("\n📂 输出:")
    print("  - companies_final.xlsx")
    print("  - aliases_final.xlsx")
    print("  - cleanup_report.md")


if __name__ == "__main__":
    main()
