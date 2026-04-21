"""
疑似重复公司检测。

从 companies.xlsx 读清洗后的公司，用多种规则识别"疑似同一家公司":
  规则 A: 去除全/半角括号后名称相同 → 99% 是重复
  规则 B: 去除括号+去除"市/县/区"后相同 → 高度可疑
  规则 C: 税号相同 → 100% 是重复（税号唯一）
  规则 D: 去除所有后缀后核心名相同，且同省 → 中度可疑（需人工判断）

输出 duplicates.xlsx: 建议合并清单（含推荐保留哪条、依据是什么）
"""
import re
from openpyxl import load_workbook, Workbook
from collections import defaultdict

INPUT = "/home/claude/work/companies.xlsx"

# 和主脚本共用的后缀词典
SUFFIXES = [
    "股份有限公司", "有限责任公司", "有限公司", "股份公司",
    "集团有限公司", "集团公司", "集团",
    "合伙企业", "普通合伙", "有限合伙",
    "个人独资企业", "个体工商户",
    "分公司", "办事处",
    "（普通合伙）", "(普通合伙)",
    "（法人独资）", "(法人独资)",
]

BUSINESS_SUFFIXES = [
    "科技", "实业", "贸易", "商贸", "工贸", "发展",
    "塑业", "塑胶", "塑化", "新材料", "材料", "化工",
    "包装", "制品", "饰品", "玩具", "食品", "饮品",
    "橡塑", "印务", "印刷", "建材",
    "进出口", "国际贸易", "国际",
    "投资", "企业管理", "管理",
]


def normalize_parens(name: str) -> str:
    """全角括号 → 半角，去除所有括号内容和空格。"""
    name = name.replace("（", "(").replace("）", ")")
    name = re.sub(r"\([^)]*\)", "", name)
    name = name.replace(" ", "").strip()
    return name


def normalize_admin_level(name: str) -> str:
    """去除行政区划字 (市/县/区/省)，用于进一步归一化。"""
    result = name
    for ch in ["市", "县", "区", "省", "自治区", "壮族", "回族", "维吾尔"]:
        result = result.replace(ch, "")
    return result


def strip_all_suffixes(name: str) -> str:
    """反复剥除所有后缀，得到最核心名。"""
    all_suf = SUFFIXES + BUSINESS_SUFFIXES
    for _ in range(5):
        new = name
        for suf in sorted(all_suf, key=len, reverse=True):
            if new.endswith(suf):
                new = new[: -len(suf)]
                break
        if new == name or len(new) < 2:
            break
        name = new
    return name


def main():
    wb = load_workbook(INPUT, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    col_name = header.index("formal_name")
    col_tax = header.index("tax_id")

    companies = []
    for i, row in enumerate(rows[1:], start=2):
        formal = (row[col_name] or "").strip()
        tax = (row[col_tax] or "").strip() if row[col_tax] else ""
        if not formal:
            continue
        companies.append({
            "row": i,
            "formal_name": formal,
            "tax_id": tax,
            "normalized_parens": normalize_parens(formal),
            "normalized_admin": normalize_admin_level(normalize_parens(formal)),
            "core": strip_all_suffixes(normalize_parens(formal)),
        })

    # 分组
    groups = []  # [(reason, confidence, [company_dict, ...])]

    # 规则 C: 税号相同 (最强信号)
    by_tax = defaultdict(list)
    for c in companies:
        if c["tax_id"] and len(c["tax_id"]) >= 15:  # 税号至少 15 位才可信
            by_tax[c["tax_id"]].append(c)
    for tax, group in by_tax.items():
        if len(group) >= 2:
            groups.append(("税号完全相同", "确定", group))

    # 标记已进入高置信组的公司，避免重复匹配
    marked = set()
    for _, _, group in groups:
        for c in group:
            marked.add(c["row"])

    # 规则 A: 去括号后完全相同
    by_parens = defaultdict(list)
    for c in companies:
        if c["row"] in marked:
            continue
        by_parens[c["normalized_parens"]].append(c)
    for key, group in by_parens.items():
        if len(group) >= 2:
            groups.append(("去除括号后名称相同", "高", group))
            for c in group:
                marked.add(c["row"])

    # 规则 B: 去括号+去行政区后相同
    by_admin = defaultdict(list)
    for c in companies:
        if c["row"] in marked:
            continue
        by_admin[c["normalized_admin"]].append(c)
    for key, group in by_admin.items():
        if len(group) >= 2 and len(key) >= 4:  # 避免短字符串误配
            groups.append(("去括号+去市县区后相同", "中高", group))
            for c in group:
                marked.add(c["row"])

    # 规则 D: 剥除所有业务后缀后核心名相同
    by_core = defaultdict(list)
    for c in companies:
        if c["row"] in marked:
            continue
        if len(c["core"]) >= 3:  # 核心名至少 3 字
            by_core[c["core"]].append(c)
    for key, group in by_core.items():
        if len(group) >= 2:
            groups.append(("剥除全部后缀后核心名相同（需人工判断）", "中", group))
            for c in group:
                marked.add(c["row"])

    # 按置信度排序: 确定 > 高 > 中高 > 中
    confidence_order = {"确定": 0, "高": 1, "中高": 2, "中": 3}
    groups.sort(key=lambda g: (confidence_order[g[1]], -len(g[2])))

    # 输出
    out = Workbook()
    ws_out = out.active
    ws_out.title = "疑似重复"
    ws_out.append([
        "组号", "置信度", "判断依据",
        "行号", "公司全名", "税号",
        "建议保留", "建议原因",
    ])

    total_dupes = 0
    for group_idx, (reason, confidence, group) in enumerate(groups, 1):
        total_dupes += len(group) - 1  # 每组 N 条，合并后剩 1 条，减 N-1

        # 推荐保留哪条：
        #  - 优先保留有税号的
        #  - 税号都有/都没有时，保留名称更长的（信息量大）
        #  - 长度相同时，保留第一次出现的
        def rank(c):
            return (
                0 if c["tax_id"] else 1,       # 有税号的排前面
                -len(c["formal_name"]),         # 长的排前面
                c["row"],                        # 行号小的排前面
            )
        sorted_group = sorted(group, key=rank)
        keeper = sorted_group[0]

        for c in sorted_group:
            is_keeper = c["row"] == keeper["row"]
            if is_keeper:
                if keeper["tax_id"]:
                    keep_reason = "✅ 保留（有税号）"
                else:
                    keep_reason = "✅ 保留（名称最长）"
            else:
                keep_reason = "❌ 删除"

            ws_out.append([
                group_idx, confidence, reason,
                c["row"], c["formal_name"], c["tax_id"],
                "保留" if is_keeper else "删除",
                keep_reason,
            ])
        # 空行分隔
        ws_out.append([])

    out.save("/home/claude/work/duplicates.xlsx")

    # 统计
    by_confidence = defaultdict(int)
    for reason, conf, group in groups:
        by_confidence[conf] += len(group) - 1

    print(f"✅ 检测完成")
    print(f"  公司总数: {len(companies)}")
    print(f"  疑似重复组数: {len(groups)}")
    print(f"  建议合并条数（可减少的记录）: {total_dupes}")
    print()
    print("  按置信度:")
    for conf in ["确定", "高", "中高", "中"]:
        if by_confidence[conf] > 0:
            print(f"    {conf}: 可合并 {by_confidence[conf]} 条")
    print()
    print("📂 输出: duplicates.xlsx")


if __name__ == "__main__":
    main()
