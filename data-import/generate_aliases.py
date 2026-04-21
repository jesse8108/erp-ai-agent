"""
根据公司全名自动生成简称。
输入: 客户档案 xlsx
输出:
  1. aliases.xlsx  (两列: 全名 | 简称，每个简称一行)
  2. companies.xlsx (清洗后的公司主表)
  3. review.md     (异常/存疑清单，给人工 review)
"""
import re
from openpyxl import load_workbook, Workbook
from collections import defaultdict, Counter

INPUT = "/home/claude/work/clients.xlsx"

# ============================================================
# 常见的"结构性后缀"和"地理前缀"——这是简称生成的核心词典
# ============================================================

# 公司性质后缀（从全名末尾剥除）
SUFFIXES = [
    "股份有限公司", "有限责任公司", "有限公司", "股份公司",
    "集团有限公司", "集团公司", "集团",
    "合伙企业", "普通合伙", "有限合伙",
    "个人独资企业", "个体工商户",
    "分公司", "办事处",
    "（普通合伙）", "(普通合伙)",
    "（法人独资）", "(法人独资)",
]

# 业务类型后缀（剥除后更口语化）
BUSINESS_SUFFIXES = [
    "科技", "实业", "贸易", "商贸", "工贸", "发展",
    "塑业", "塑胶", "塑化", "新材料", "材料", "化工",
    "包装", "制品", "饰品", "玩具", "食品", "饮品",
    "橡塑", "印务", "印刷", "建材",
    "进出口", "国际贸易", "国际",
    "投资", "企业管理", "管理",
]

# 省级行政区
PROVINCES = [
    "北京", "天津", "上海", "重庆",
    "河北", "山西", "辽宁", "吉林", "黑龙江",
    "江苏", "浙江", "安徽", "福建", "江西", "山东",
    "河南", "湖北", "湖南", "广东", "海南",
    "四川", "贵州", "云南", "陕西", "甘肃", "青海",
    "内蒙古", "广西", "西藏", "宁夏", "新疆",
    "台湾", "香港", "澳门",
]

# 常见地级市前缀（只列样本里出现过的高频项，不全面但够用）
# 这个列表是可扩展的，遇到没覆盖的字样会走通用规则
CITY_PREFIXES = [
    # 江苏
    "南京", "苏州", "无锡", "常州", "南通", "扬州", "镇江", "泰州", "徐州", "连云港", "盐城",
    "淮安", "宿迁", "张家港", "江阴", "昆山", "常熟", "太仓", "宜兴",
    # 浙江
    "杭州", "宁波", "温州", "嘉兴", "湖州", "绍兴", "金华", "衢州", "舟山", "台州", "丽水",
    "义乌", "慈溪", "余姚", "诸暨", "桐乡", "海宁", "永康", "乐清", "平湖", "德清",
    # 广东
    "广州", "深圳", "东莞", "佛山", "中山", "珠海", "汕头", "惠州", "江门", "湛江", "肇庆",
    "茂名", "梅州", "汕尾", "河源", "阳江", "清远", "潮州", "揭阳", "云浮",
    # 山东
    "济南", "青岛", "淄博", "枣庄", "东营", "烟台", "潍坊", "济宁", "泰安", "威海",
    "日照", "临沂", "德州", "聊城", "滨州", "菏泽",
    # 福建
    "福州", "厦门", "泉州", "漳州", "莆田", "三明", "南平", "龙岩", "宁德", "晋江",
    # 安徽
    "合肥", "芜湖", "蚌埠", "淮南", "马鞍山", "淮北", "铜陵", "安庆", "黄山", "滁州",
    "阜阳", "宿州", "六安", "亳州", "池州", "宣城", "桐城",
    # 河南
    "郑州", "开封", "洛阳", "平顶山", "安阳", "鹤壁", "新乡", "焦作", "濮阳", "许昌",
    "漯河", "三门峡", "南阳", "商丘", "信阳", "周口", "驻马店",
    # 河北
    "石家庄", "唐山", "秦皇岛", "邯郸", "邢台", "保定", "张家口", "承德", "沧州", "廊坊",
    "衡水", "辛集",
    # 湖北
    "武汉", "黄石", "十堰", "宜昌", "襄阳", "鄂州", "荆门", "孝感", "荆州", "黄冈",
    "咸宁", "随州", "恩施",
    # 湖南
    "长沙", "株洲", "湘潭", "衡阳", "邵阳", "岳阳", "常德", "张家界", "益阳", "郴州",
    "永州", "怀化", "娄底",
    # 四川
    "成都", "绵阳", "德阳", "自贡", "攀枝花", "泸州", "广元", "遂宁", "内江", "乐山",
    "南充", "眉山", "宜宾", "广安", "达州", "雅安", "资阳", "邛崃",
    # 辽宁
    "沈阳", "大连", "鞍山", "抚顺", "本溪", "丹东", "锦州", "营口", "阜新", "辽阳",
    "盘锦", "铁岭", "朝阳", "葫芦岛",
    # 江西
    "南昌", "景德镇", "萍乡", "九江", "新余", "鹰潭", "赣州", "吉安", "宜春", "抚州", "上饶",
    # 山西
    "太原", "大同", "阳泉", "长治", "晋城", "朔州", "晋中", "运城", "忻州", "临汾", "吕梁",
    # 陕西
    "西安", "铜川", "宝鸡", "咸阳", "渭南", "延安", "汉中", "榆林", "安康", "商洛",
    # 广西
    "南宁", "柳州", "桂林", "梧州", "北海", "防城港", "钦州", "贵港", "玉林", "百色",
    "贺州", "河池", "来宾", "崇左",
    # 云南
    "昆明", "曲靖", "玉溪", "保山", "昭通", "丽江", "普洱", "临沧",
    # 贵州
    "贵阳", "六盘水", "遵义", "安顺", "毕节", "铜仁",
    # 海南
    "海口", "三亚", "儋州",
    # 其他省会/直辖市/较高频
    "呼和浩特", "银川", "乌鲁木齐", "兰州", "西宁", "拉萨",
    "长春", "哈尔滨",
]


def strip_suffix(name: str, suffixes: list[str]) -> str:
    """从末尾剥除一个后缀（匹配最长的）。"""
    for suf in sorted(suffixes, key=len, reverse=True):
        if name.endswith(suf):
            return name[: -len(suf)]
    return name


def strip_suffix_repeatedly(name: str, suffixes: list[str], max_rounds: int = 3) -> str:
    """
    反复剥除业务后缀，处理"新材料科技"这种复合情况。
    例如：辉凡新材料科技 → 辉凡新材料 → 辉凡（剥了 2 次）
    """
    for _ in range(max_rounds):
        stripped = strip_suffix(name, suffixes)
        if stripped == name or len(stripped) < 2:
            break
        name = stripped
    return name


def strip_parentheses(name: str) -> str:
    """去除公司名中的括号内容，如 '涂多多（青岛）跨境电子商务' → '涂多多跨境电子商务'。"""
    # 中英文括号都处理
    name = re.sub(r"[（(][^）)]*[）)]", "", name)
    return name.strip()


def strip_city_prefix(name: str) -> tuple[str, str | None]:
    """
    从开头剥除地理前缀（省、市）。
    返回 (剥除后的名称, 被剥掉的前缀或None)。
    支持形式：
      - 浙江XX / 浙江省XX
      - 张家港XX / 张家港市XX
      - 浙江义乌XX / 浙江省义乌市XX
      - 义乌市廿三里街道XX  → 匹配到"义乌市"后停，保留廿三里街道
    """
    original = name
    prefix_parts = []

    # 先尝试省
    for province in sorted(PROVINCES, key=len, reverse=True):
        if name.startswith(province):
            # 允许带"省"、"市"、"自治区"等字
            if name.startswith(province + "省"):
                prefix_parts.append(province + "省")
                name = name[len(province) + 1 :]
            elif name.startswith(province + "市"):
                prefix_parts.append(province + "市")
                name = name[len(province) + 1 :]
            elif name.startswith(province + "自治区"):
                prefix_parts.append(province + "自治区")
                name = name[len(province) + 3 :]
            elif name.startswith(province + "壮族自治区"):
                prefix_parts.append(province + "壮族自治区")
                name = name[len(province) + 5 :]
            else:
                prefix_parts.append(province)
                name = name[len(province) :]
            break

    # 再尝试市
    for city in sorted(CITY_PREFIXES, key=len, reverse=True):
        if name.startswith(city):
            if name.startswith(city + "市"):
                prefix_parts.append(city + "市")
                name = name[len(city) + 1 :]
            else:
                prefix_parts.append(city)
                name = name[len(city) :]
            break

    if not prefix_parts:
        return original, None
    return name, "".join(prefix_parts)


def generate_aliases(formal_name: str) -> list[str]:
    """
    根据全名生成简称列表。
    策略：
      0. 先去除括号内容（如"涂多多（青岛）" → "涂多多"）
      1. 剥除公司性质后缀 → 得到"核心名"
      2. 从核心名里剥除地理前缀 → 得到"品牌+业务"
      3. 反复剥除业务类型后缀 → 得到"最短口语名"
      4. 生成多个变体
    """
    aliases = set()

    # Step 0: 去括号
    cleaned = strip_parentheses(formal_name)
    if not cleaned:
        return []

    # Step 1: 剥除公司性质（反复剥，处理"股份有限公司" + 尾部还有"集团"等）
    core = strip_suffix_repeatedly(cleaned, SUFFIXES).strip()
    if not core:
        return []

    # Step 2: 剥离地理前缀
    without_geo, geo_prefix = strip_city_prefix(core)

    # Step 3: 反复剥业务后缀（"新材料科技" → "新材料" → 核心词）
    shortest = strip_suffix_repeatedly(without_geo, BUSINESS_SUFFIXES).strip()

    # 中间形态：只剥一次业务后缀的版本（如"辉凡新材料"）
    mid = strip_suffix(without_geo, BUSINESS_SUFFIXES).strip()

    # 生成候选
    candidates = []

    # 最短口语名（如"辉凡"）
    if shortest and len(shortest) >= 2:
        candidates.append(shortest)

    # 中间形态（如"辉凡新材料"）——只有当它和 shortest 不同时才加
    if mid and mid != shortest and len(mid) >= 2:
        candidates.append(mid)

    # 地理 + 最短口语名（如"张家港辉凡"）
    if geo_prefix and shortest and len(shortest) >= 2:
        clean_geo = geo_prefix.replace("省", "").replace("市", "").replace("自治区", "").replace("壮族", "")
        if clean_geo:
            candidates.append(clean_geo + shortest)

    # 去重 + 过滤
    for alias in candidates:
        alias = alias.strip()
        if not alias or len(alias) < 2:
            continue
        if alias == formal_name or alias == cleaned:
            continue
        aliases.add(alias)

    return sorted(aliases, key=len)


def is_problematic(formal_name: str, aliases: list[str]) -> str | None:
    """
    检测异常，返回原因（None 表示正常）。
    """
    if not formal_name or not formal_name.strip():
        return "公司名为空"
    if not aliases:
        return "未能生成任何简称（可能全名结构特殊）"
    # 太短的全名（不含后缀）可能是已经是简称了
    core = strip_suffix(formal_name, SUFFIXES)
    if len(core) <= 3:
        return f"剥除后缀后只剩 {len(core)} 字，可能是不规范录入"
    return None


# ============================================================
# 主流程
# ============================================================

def main():
    wb = load_workbook(INPUT, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data_rows = rows[1:]

    # 字段索引
    COL_ID = header.index("编号")
    COL_TYPE = header.index("类型")
    COL_NAME = header.index("公司名称")
    COL_TAX = header.index("纳税人识别号")
    COL_STATUS = header.index("经营状态")
    COL_LEGAL = header.index("法定代表人")
    COL_SUBMITTER = header.index("提交人")

    # 收集结果
    companies = []      # 清洗后的公司主表
    alias_rows = []     # (全名, 简称)
    problems = []       # 存疑清单
    alias_conflicts = defaultdict(list)  # alias → [公司名, ...] 用于检测冲突

    seen_names = {}  # 去重：formal_name → 首次出现的行号
    duplicates = []

    for i, row in enumerate(data_rows, start=2):
        formal_name = (row[COL_NAME] or "").strip()
        if not formal_name:
            problems.append({"row": i, "issue": "公司名为空", "name": "", "aliases": []})
            continue

        # 去重（以全名为准）
        if formal_name in seen_names:
            duplicates.append({"row": i, "first_row": seen_names[formal_name], "name": formal_name})
            continue
        seen_names[formal_name] = i

        # 生成简称
        aliases = generate_aliases(formal_name)

        # 异常检测
        issue = is_problematic(formal_name, aliases)
        if issue:
            problems.append({"row": i, "issue": issue, "name": formal_name, "aliases": aliases})

        # 冲突检测（同一个简称指向多家公司）
        for alias in aliases:
            alias_conflicts[alias].append(formal_name)

        # 写入结果
        companies.append({
            "formal_name": formal_name,
            "tax_id": (row[COL_TAX] or "").strip() if row[COL_TAX] else "",
            "type_raw": (row[COL_TYPE] or "").strip() if row[COL_TYPE] else "",
            "status_raw": (row[COL_STATUS] or "").strip() if row[COL_STATUS] else "",
            "legal_person": (row[COL_LEGAL] or "").strip() if row[COL_LEGAL] else "",
            "submitter": (row[COL_SUBMITTER] or "").strip() if row[COL_SUBMITTER] else "",
            "external_id": (row[COL_ID] or "").strip() if row[COL_ID] else "",
        })

        for alias in aliases:
            alias_rows.append((formal_name, alias))

    # 识别真正的冲突（同一简称对应 2+ 个公司）
    real_conflicts = {a: names for a, names in alias_conflicts.items() if len(names) > 1}

    # ============================================================
    # 输出 1: aliases.xlsx
    # ============================================================
    out_aliases = Workbook()
    ws_a = out_aliases.active
    ws_a.title = "aliases"
    ws_a.append(["正式名", "简称"])
    for formal, alias in alias_rows:
        ws_a.append([formal, alias])
    out_aliases.save("/home/claude/work/aliases.xlsx")

    # ============================================================
    # 输出 2: companies.xlsx (清洗后的主表)
    # ============================================================
    out_companies = Workbook()
    ws_c = out_companies.active
    ws_c.title = "companies"
    ws_c.append([
        "formal_name", "tax_id", "type_raw", "status_raw",
        "legal_person", "submitter", "external_id",
    ])
    for c in companies:
        ws_c.append([
            c["formal_name"], c["tax_id"], c["type_raw"], c["status_raw"],
            c["legal_person"], c["submitter"], c["external_id"],
        ])
    out_companies.save("/home/claude/work/companies.xlsx")

    # ============================================================
    # 输出 3: review.md
    # ============================================================
    with open("/home/claude/work/review.md", "w", encoding="utf-8") as f:
        f.write("# 客户档案清洗 - 人工 Review 清单\n\n")
        f.write(f"- 总行数（不含表头）: {len(data_rows)}\n")
        f.write(f"- 成功清洗: {len(companies)}\n")
        f.write(f"- 空行/异常: {len(problems)}\n")
        f.write(f"- 重复（按全名去重）: {len(duplicates)}\n")
        f.write(f"- 生成的简称总数: {len(alias_rows)}\n")
        f.write(f"- 平均每公司 {len(alias_rows) / max(len(companies), 1):.1f} 个简称\n\n")

        f.write("## 1. 简称冲突（同一简称指向多家公司）\n\n")
        f.write("这些简称需要业务员确认：是让 AI 反问用户选择，还是删掉其中一些。\n\n")
        if real_conflicts:
            f.write(f"共 {len(real_conflicts)} 个冲突简称：\n\n")
            f.write("| 简称 | 对应公司 |\n|------|----------|\n")
            for alias, names in sorted(real_conflicts.items(), key=lambda x: -len(x[1]))[:50]:
                f.write(f"| {alias} | {' / '.join(names)} |\n")
            if len(real_conflicts) > 50:
                f.write(f"\n...还有 {len(real_conflicts) - 50} 个，见 conflicts.xlsx\n")
        else:
            f.write("无冲突 ✅\n")
        f.write("\n")

        f.write("## 2. 重复的公司名（已自动去重，保留首次出现的行）\n\n")
        if duplicates:
            f.write(f"共 {len(duplicates)} 条重复：\n\n")
            f.write("| 行号 | 重复的公司名 | 首次出现行 |\n|------|-------------|------------|\n")
            for d in duplicates[:30]:
                f.write(f"| {d['row']} | {d['name']} | {d['first_row']} |\n")
            if len(duplicates) > 30:
                f.write(f"\n...还有 {len(duplicates) - 30} 条\n")
        else:
            f.write("无重复 ✅\n")
        f.write("\n")

        f.write("## 3. 未能生成简称的公司\n\n")
        no_alias = [p for p in problems if "未能生成" in p["issue"]]
        if no_alias:
            f.write(f"共 {len(no_alias)} 条，通常是全名太短或结构特殊：\n\n")
            for p in no_alias[:30]:
                f.write(f"- 行 {p['row']}: `{p['name']}`\n")
        else:
            f.write("全部成功生成简称 ✅\n")
        f.write("\n")

        f.write("## 4. 其他异常\n\n")
        other = [p for p in problems if "未能生成" not in p["issue"]]
        if other:
            for p in other[:30]:
                f.write(f"- 行 {p['row']}: {p['issue']} — `{p['name']}`\n")
        else:
            f.write("无 ✅\n")

    # ============================================================
    # 冲突清单导出
    # ============================================================
    if real_conflicts:
        out_conflicts = Workbook()
        ws_x = out_conflicts.active
        ws_x.title = "conflicts"
        ws_x.append(["简称", "对应公司数", "对应的公司全名"])
        for alias, names in sorted(real_conflicts.items(), key=lambda x: -len(x[1])):
            ws_x.append([alias, len(names), " / ".join(names)])
        out_conflicts.save("/home/claude/work/conflicts.xlsx")

    # 控制台摘要
    print(f"✅ 处理完成")
    print(f"  公司数: {len(companies)}")
    print(f"  简称数: {len(alias_rows)}（平均 {len(alias_rows)/max(len(companies),1):.1f}/家）")
    print(f"  冲突简称: {len(real_conflicts)}")
    print(f"  重复: {len(duplicates)}")
    print(f"  异常: {len(problems)}")
    print()
    print("📂 输出文件:")
    print("  - companies.xlsx  (清洗后的公司主表)")
    print("  - aliases.xlsx    (正式名-简称 对应表)")
    print("  - conflicts.xlsx  (简称冲突清单)")
    print("  - review.md       (人工 review 清单)")


if __name__ == "__main__":
    main()
